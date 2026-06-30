"""Write GCC index last-prices to 'GCC Indices.xlsx' — all sources run locally.

  Yahoo:       TASI, DFM
  Investing:   ADX, QE All Share, Kuwait All Share
  TradingView: Bahrain All Share
  Manual:      Oman MSMTR, S&P GCC  (manual_indices.json — no free source; preserved)

Colour: green = live, amber = manual, red = unavailable.
"""
from __future__ import annotations
import json
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font
import yfinance as yf

import indices_config as IC
from index_sources import investing_last, tv_last

HERE = Path(__file__).resolve().parent
OUT = HERE / "GCC Indices.xlsx"
MANUAL = HERE / "manual_indices.json"

GREEN = PatternFill("solid", fgColor="C6EFCE")
AMBER = PatternFill("solid", fgColor="FFEB9C")
RED   = PatternFill("solid", fgColor="FFC7CE")


def _yahoo_last(sym: str):
    try:
        c = yf.Ticker(sym).history(period="5d")["Close"].dropna()
        return round(float(c.iloc[-1]), 2) if not c.empty else None
    except Exception:
        return None


def _sane(v, bbg):
    ref = IC.REFERENCE.get(bbg)
    return v is not None and ref and 0.5 * ref <= v <= 2 * ref


def main():
    manual = json.loads(MANUAL.read_text()) if MANUAL.exists() else {}
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Indices"
    ws.append(["Bloomberg Ticker", "Index", "Market", "Last Price", "As of", "Source"])
    for c in ws[1]:
        c.font = Font(bold=True)

    for bbg, (name, market, ysym, slug, tv) in IC.INDICES.items():
        last = asof = None; source = "unavailable"; fill = RED
        if ysym:
            last = _yahoo_last(ysym)
            if last is not None: source, asof, fill = "Yahoo", today, GREEN
        elif slug:
            v, _ = investing_last(slug)
            if _sane(v, bbg): last, source, asof, fill = round(v, 2), "Investing", today, GREEN
        elif tv:
            v = tv_last(tv)
            if _sane(v, bbg): last, source, asof, fill = round(v, 2), "TradingView", today, GREEN
        if last is None and bbg in manual and isinstance(manual[bbg], dict):
            last = manual[bbg].get("last"); asof = manual[bbg].get("asof")
            if last is not None: source, fill = "Manual (Bloomberg)", AMBER

        ws.append([bbg, name, market, last, asof, source])
        ws.cell(row=ws.max_row, column=4).fill = fill

    for col, w in zip("ABCDEF", (16, 26, 18, 12, 12, 18)):
        ws.column_dimensions[col].width = w
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws.append([]); ws.append([f"Last refreshed: {stamp}"])
    wb.save(OUT)
    priced = sum(1 for r in range(2, 2 + len(IC.INDICES)) if ws.cell(row=r, column=4).value is not None)
    live = sum(1 for r in range(2, 2 + len(IC.INDICES))
               if ws.cell(row=r, column=6).value in ("Yahoo", "Investing", "TradingView"))
    print(f"Wrote {OUT}  ({priced}/{len(IC.INDICES)} priced, {live} live)")


if __name__ == "__main__":
    main()
