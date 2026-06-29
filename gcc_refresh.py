"""GCC sheet refresh — Phase 1 (market + ratios + reported actuals).

Reuses the final_earnings engine (src/providers/yahoo.py) as the data layer.
Writes ONLY the scraped columns (G-AN) into a COPY of the workbook, with
per-cell provenance colouring. Columns A-F are never touched.

Run with the engine's venv so `import src...` resolves:
    PYTHONPATH=/Users/rabigaarip/final_earnings \
    /Users/rabigaarip/final_earnings/.venv/bin/python gcc_refresh.py
"""
from __future__ import annotations
import shutil, sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import column_index_from_string

from src.providers.yahoo import fetch_quote, fetch_financials  # engine reuse
from src.providers._yahoo_blind import is_yahoo_blind

import column_map as CM
from ticker_map import MAP as TICKER_MAP
from ms_forecast import forecast_for

# ── provenance colours ────────────────────────────────────────
GREEN  = PatternFill("solid", fgColor="C6EFCE")  # live actual
AMBER  = PatternFill("solid", fgColor="FFEB9C")  # computed / stale
RED    = PatternFill("solid", fgColor="FFC7CE")  # missing — needs fallback/manual
BLUE   = PatternFill("solid", fgColor="DDEBF7")  # estimate pending (Phase 4)

import os
HERE = Path(__file__).resolve().parent
SRC_XLSX = Path(os.environ.get("GCC_MASTER_XLSX",
                Path.home() / "Downloads" / "Stock Financial Data - GCC.xlsx"))
OUT_XLSX = HERE / "Stock Financial Data - GCC (populated).xlsx"

# Bloomberg -> Yahoo resolution: full 60-name map (verified). Blind UAE/Oman
# names resolve to a symbol but are flagged for the Phase-3 fallback.
BBG_TO_YAHOO = {bbg: yh for bbg, (yh, _blind) in TICKER_MAP.items()}
BLIND = {bbg for bbg, (_yh, blind) in TICKER_MAP.items() if blind}


def _cell(ws, col_letter: str, row: int):
    return ws.cell(row=row, column=column_index_from_string(col_letter))

def _write(ws, col, row, value, fill):
    c = _cell(ws, col, row)
    c.value = value
    c.fill = fill

def _period_index(periods):
    return {p.period_label: p for p in periods}

def _history_value(quote, kind: str):
    """Return (value, fill) for H/I/J from the 1Y price history arrays."""
    dates, prices = quote.price_history_dates, quote.price_history_prices
    if not dates:
        return None, RED
    pairs = list(zip(dates, prices))
    def close_on_or_before(iso):
        elig = [p for d, p in pairs if d <= iso]
        return elig[-1] if elig else None
    if kind == "2025-12-31":
        v = close_on_or_before("2025-12-31")
        return (round(v, 4) if v else None), (GREEN if v else RED)
    if kind == "ytd":
        base, cur = close_on_or_before("2025-12-31"), prices[-1]
        return ((round((cur - base) / base, 4), AMBER) if base else (None, RED))
    if kind == "qtd":
        base, cur = close_on_or_before("2026-03-31"), prices[-1]
        return ((round((cur - base) / base, 4), AMBER) if base else (None, RED))
    return None, RED


def _write_ms(ws, row: int, yh: str, blind: bool):
    """Fill estimate columns (annual + quarterly) from the MS forecast snapshot,
    and backfill any empty actual cells (bank gaps, or everything for blind names)."""
    fc = forecast_for(yh)
    ann, qtr = fc.get("annual", {}), fc.get("quarterly", {})

    def val(label, field):
        src = ann if label.startswith("FY") else qtr
        d = src.get(label)
        return d.get(field) if d else None

    # Estimates — MS-owned columns
    for col, lbl in CM.REVENUE_ESTIMATE_COLS.items():
        v = val(lbl, "revenue")
        _write(ws, col, row, round(v, 2) if v is not None else None, GREEN if v is not None else BLUE)
    for col, lbl in CM.NETPROFIT_ESTIMATE_COLS.items():
        v = val(lbl, "net_income")
        _write(ws, col, row, round(v, 2) if v is not None else None, GREEN if v is not None else BLUE)

    # Actuals backfill — only fill cells Yahoo left empty (never overwrite a green)
    for col, lbl in CM.REVENUE_ACTUAL_COLS.items():
        if _cell(ws, col, row).value is None:
            v = val(lbl, "revenue")
            if v is not None:   _write(ws, col, row, round(v, 2), AMBER)
            elif blind:         _write(ws, col, row, None, RED)
    for col, lbl in CM.NETPROFIT_ACTUAL_COLS.items():
        if _cell(ws, col, row).value is None:
            v = val(lbl, "net_income")
            if v is not None:   _write(ws, col, row, round(v, 2), AMBER)
            elif blind:         _write(ws, col, row, None, RED)


def refresh_row(ws, row: int, bbg: str):
    yh = BBG_TO_YAHOO.get(bbg)
    name = ws.cell(row=row, column=2).value
    currency = ws.cell(row=row, column=3).value or "USD"
    if not yh:
        print(f"  row {row:<3} {bbg:<18} UNRESOLVED — no Yahoo symbol")
        return
    blind = (bbg in BLIND) or is_yahoo_blind(yh)
    print(f"  row {row:<3} {bbg:<18} -> {yh:<10} {'(BLIND)' if blind else '       '} {name}")
    if blind:
        # Yahoo has no price for blind names; MS forecast supplies actuals + estimates.
        for col in list(CM.QUOTE_COLS) + list(CM.HISTORY_COLS):
            _write(ws, col, row, None, RED)
        _write_ms(ws, row, yh, blind=True)
        return

    # 1) Market + ratios
    q = fetch_quote(yh)
    if q:
        _write(ws, "G",  row, round(q.price, 4) if q.price else None, GREEN if q.price else RED)
        _write(ws, "AK", row, round(q.trailing_pe, 2) if q.trailing_pe else None, GREEN if q.trailing_pe else RED)
        _write(ws, "AL", row, round(q.price_to_book, 2) if q.price_to_book else None, GREEN if q.price_to_book else RED)
        dy = q.dividend_yield
        _write(ws, "AM", row, round(dy, 2) if dy else None, GREEN if dy else RED)  # already a % in this feed
        dps = round((dy/100.0) * q.price, 4) if (dy and q.price) else None
        _write(ws, "AN", row, dps, AMBER if dps else RED)
        for col, (_lbl, kind) in CM.HISTORY_COLS.items():
            v, fill = _history_value(q, kind)
            _write(ws, col, row, v, fill)
    else:
        for col in list(CM.QUOTE_COLS) + list(CM.HISTORY_COLS):
            _write(ws, col, row, None, RED)

    # 2) Reported actuals
    fin = fetch_financials(yh, currency, is_bank=False)
    qidx = _period_index(fin["quarterly"]); aidx = _period_index(fin["annual"])
    for col, plabel in CM.REVENUE_ACTUAL_COLS.items():
        src = aidx if plabel.startswith("FY") else qidx
        p = src.get(plabel)
        val = p.revenue if p else None
        _write(ws, col, row, round(val, 2) if val else None, GREEN if val else RED)
    for col, plabel in CM.NETPROFIT_ACTUAL_COLS.items():
        src = aidx if plabel.startswith("FY") else qidx
        p = src.get(plabel)
        val = p.net_income if p else None
        _write(ws, col, row, round(val, 2) if val else None, GREEN if val else RED)

    # 3) MarketScreener forecast snapshot — fills estimates (annual + quarterly)
    #    AND backfills any actual cells Yahoo left blank (bank gaps).
    _write_ms(ws, row, yh, blind=False)


def main(rows=None):
    shutil.copyfile(SRC_XLSX, OUT_XLSX)
    wb = openpyxl.load_workbook(OUT_XLSX)
    ws = wb["Sheet1"]
    if rows is None:
        rows = [r for r in range(3, ws.max_row + 1)
                if ws.cell(row=r, column=1).value in BBG_TO_YAHOO]
    print(f"Refreshing {len(rows)} row(s):")
    for r in rows:
        refresh_row(ws, r, ws.cell(row=r, column=1).value)
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    note = ws.cell(row=1, column=1); note.value = f"Last refreshed: {stamp}"
    note.font = Font(italic=True, size=9)
    wb.save(OUT_XLSX)
    print(f"\nSaved -> {OUT_XLSX}")


if __name__ == "__main__":
    main()
